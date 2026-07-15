# -*- coding: utf-8 -*-
import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SUFFIX = (
    "Warm natural window light, realistic food photography, Korean home kitchen "
    "or small bakery classroom, shallow depth of field, no people faces, no text, "
    "no logos, no watermark, 16:9 aspect ratio"
)

PROMPTS = {
    "why-baker-certification/childhood-bread": "Nostalgic Korean-style chestnut loaf bread in a small neighborhood bakery display case, golden brown top with visible chestnut pieces, soft nostalgic mood, childhood memory feeling",
    "why-baker-certification/after-pass": "Freshly baked plain milk loaf on a wooden cutting board next to an open notebook, sense of new chapter beginning after exam, calm hopeful mood",
    "baker-cert-series-roadmap/series-purpose": "Wide shot of a small bakery workshop with several loaves on racks, organized learning atmosphere, roadmap and series planning mood",
    "baker-cert-series-roadmap/how-to-read": "Open cookbook and handwritten reading order notes beside a simple loaf, top-down desk shot, study sequence planning mood",
    "baker-cert-8month-roadmap/exam-structure": "Clean desk with exam prep checklist, written and practical exam structure sketched on paper (no readable text), highlighters and timer",
    "baker-cert-8month-roadmap/if-again": "Baker practicing at home counter, multiple trial loaves of different quality side by side, reflection and improvement mood",
    "baker-cert-practical-mistakes/intro": "Bakery classroom with student practice station, dough and tools ready, long practice journey mood",
    "baker-cert-practical-mistakes/dough-failures": "Two dough balls same size but different surface texture, digital scale nearby, comparison of wrong hydration result",
    "baker-cert-practical-mistakes/fermentation-failures": "Over-proofed and under-proofed dough in bowls side by side, timer in background, fermentation mistake comparison",
    "baker-cert-practical-mistakes/shaping-failures": "Misshapen loaf before baking, uneven surface, correct weight on scale but poor shaping",
    "baker-cert-practical-mistakes/baking-failures": "Home oven with unevenly baked loaf, one side too dark one side pale, oven door slightly open steam visible",
    "baker-cert-practical-mistakes/time-practice": "Baker hands shaping dough with kitchen timer prominently placed, speed practice focus not taste",
    "baker-cert-written-tips/intro": "Desk piled with theory textbooks and sticky notes, procrastination mood, written exam study postponed",
    "baker-cert-written-tips/materials-nutrition": "Flour bags, yeast, butter, eggs arranged in grouped clusters on counter, ingredient study layout",
    "baker-cert-exam-day-pass/intro": "Early morning prep scene, exam day bag and apron ready by door, documentary record mood",
    "baker-cert-exam-day-pass/pass-moment": "Hands holding phone showing pass result (screen blurred no text), relieved exhale mood, quiet celebration",
    "baker-cert-exam-day-pass/after-pass-immediate": "Notebook open immediately after exam writing quick notes, half-eaten practice roll nearby, immediate debrief mood",
    "baker-cert-to-bread-rd/intro": "Rest day mood, certification loaf cooling on rack, no new experiment started yet, pause before R&D",
    "baker-cert-to-bread-rd/night-bread-bridge": "Korean chestnut bread loaf cross-section showing soft interior and chestnut topping, project destination mood",
    "baker-cert-to-bread-rd/tools-and-log": "Baking scale, oven thermometer, pen and experiment log notebook on kitchen counter, R&D tools flat lay",
    "bread-rd-series-guide/goal": "Index cards numbered 1-9 laid out showing R&D reading order (numbers blurred), project guide mood",
    "bread-rd-series-guide/failures": "Confused reader desk with too many open tabs feeling, scattered bread photos and notes, overwhelm mood",
    "bread-rd-series-guide/one-variable": "Single line highlighted in notebook one variable at a time, minimal focused study shot",
    "bread-rd-series-guide/next": "Empty page in logbook titled next experiment, pencil ready, upcoming posts mood",
    "bread-rd-night-bread-v1/goal": "Ideal nostalgic Korean chestnut bread memory reference, glossy top dense chestnut pieces, target goal appearance",
    "bread-rd-night-bread-v1/failures": "First trial chestnut bread with separated topping, dry surface, not matching memory, visible failures",
    "bread-rd-night-bread-v1/one-variable": "Single changed variable focus: chestnut syrup jar highlighted among unchanged ingredients",
    "bread-rd-night-bread-v1/next": "Notebook page planning next trial with one arrow pointing to topping timing change",
    "bread-rd-night-bread-v1/ingredient-note": "Canned chestnuts and homemade syrup side by side on counter, ingredient selection comparison",
    "bread-rd-night-bread-v2/goal": "Chestnut bread with topping placed earlier in process, goal appearance for trial 2",
    "bread-rd-night-bread-v2/failures": "Topping sunk unevenly, some chestnut pieces fallen off edges, trial 2 failure close-up",
    "bread-rd-night-bread-v2/one-variable": "Hands placing chestnut topping earlier on proofed dough, single variable timing change",
    "bread-rd-night-bread-v2/next": "Note saying next change hydration plus 2 percent, planning shot",
    "bread-rd-night-bread-v3/goal": "Slightly more moist chestnut bread goal, softer crumb visible in cut slice",
    "bread-rd-night-bread-v3/failures": "Sticky collapsed top from too much hydration, wet shine on surface, failure state",
    "bread-rd-night-bread-v3/one-variable": "Measuring spoon adding small amount water to dough, hydration plus 2 percent single change",
    "bread-rd-night-bread-v3/next": "Note planning storage method comparison for next trial",
    "bread-rd-night-bread-v4/goal": "Same loaf quality goal, focus on overnight storage comparison setup",
    "bread-rd-night-bread-v4/failures": "Next-day slices showing dryness difference between storage methods, side by side comparison",
    "bread-rd-night-bread-v4/one-variable": "Loaf in loose bag vs open plate storage, single variable storage method",
    "bread-rd-night-bread-v4/next": "Note planning syrup simmer time plus 2 minutes",
    "bread-rd-night-bread-v5/goal": "Chestnut bread with improved syrup gloss goal appearance",
    "bread-rd-night-bread-v5/failures": "Syrup too thick causing uneven shine, some spots overly sweet glossy patches",
    "bread-rd-night-bread-v5/one-variable": "Small saucepan with chestnut syrup simmering extra 2 minutes, timer visible",
    "bread-rd-night-bread-v5/next": "Note planning reduce syrup sugar slightly for trial 6",
    "bread-rd-night-bread-mid-review/goal": "Five trial loaves labeled 1-5 in row for mid-project review comparison",
    "bread-rd-night-bread-mid-review/failures": "Collage feel of recurring problems: topping shift, moisture, storage dryness still present",
    "bread-rd-night-bread-mid-review/one-variable": "Notebook page with fixed values draft checklist for chestnut bread R&D",
    "bread-rd-night-bread-mid-review/next": "List of candidate variables for trials 6-9, forward planning mood",
    "bread-rd-night-bread-v6/goal": "Less sweet chestnut bread goal, balanced syrup sweetness",
    "bread-rd-night-bread-v6/failures": "Slightly bland surface, reduced sugar made syrup less adhesive, visible issue",
    "bread-rd-night-bread-v6/one-variable": "Sugar bowl with small reduction amount removed, syrup sugar minus single change",
    "bread-rd-night-bread-v6/next": "Fresh chestnuts on counter planning switch from canned for trial 7",
    "bread-rd-night-bread-v7/goal": "Chestnut bread with fresh peeled chestnuts goal, better aroma appearance",
    "bread-rd-night-bread-v7/failures": "Uneven chestnut piece sizes on top, lumpy appearance, fresh chestnut size variance failure",
    "bread-rd-night-bread-v7/one-variable": "Fresh boiled peeled chestnuts replacing canned, single ingredient swap focus",
    "bread-rd-night-bread-v7/next": "Note planning post-bake thin syrup brush for trial 8",
    "bread-rd-night-bread-v8/goal": "Best trial so far chestnut bread, improved gloss and adhesion goal",
    "bread-rd-night-bread-v8/failures": "Slight surface dryness on crust edges despite better gloss, minor failure detail",
    "bread-rd-night-bread-v8/one-variable": "Pastry brush applying thin syrup immediately after baking on hot loaf",
    "bread-rd-night-bread-v8/next": "Winter indoor note planning fermentation time adjustment for trial 9",
    "bread-rd-night-bread-v9/goal": "Reproduce trial 8 fixed values in winter heated room, same bread goal",
    "bread-rd-night-bread-v9/failures": "Winter trial slightly drier crust than summer trial 8, subtle dryness failure",
    "bread-rd-night-bread-v9/one-variable": "Proofing bowl with extended fermentation 58 minutes note, winter time adjustment only",
    "bread-rd-night-bread-v9/next": "Humidity and storage notes for trial 10 planning",
    "bread-rd-night-bread-practical-guide/goal": "Clean summary layout: 7 principles icons feel, chestnut bread and checklist mood, practical guide hero",
    "bread-rd-night-bread-practical-guide/memo-template": "Blank experiment log template on paper: goal, 3 failures, cause, one variable columns (no readable text)",
    "baker-cert-one-page-cheatsheet/goal": "One-page cheat sheet concept, compressed exam prep notes beside simple loaf",
    "baker-cert-one-page-cheatsheet/exam-structure": "Visual diagram feel of written plus practical exam structure on single sheet (no readable text)",
    "baker-cert-one-page-cheatsheet/exam-day": "Exam day checklist items with timer, apron, tools packed, morning preparation flat lay",
}


def load_posts():
    raw = (ROOT / "data/posts.js").read_text(encoding="utf-8")
    return json.loads(raw.replace("window.POSTS_DATA =", "", 1).strip().rstrip(";"))


def main():
    assign = json.loads((ROOT / "data/photo-assignments.json").read_text(encoding="utf-8"))["assignments"]
    posts = load_posts()
    titles = {p["slug"]: p["title"] for p in posts}

    DONE_UNTIL = 12

    rows = []
    for post in posts:
        for sec in post.get("sections", []):
            m = re.search(r"photos/([^/]+)/([^.\s]+)\.jpg", sec.get("content", ""))
            if not m:
                continue
            slug, section_id = m.group(1), m.group(2)
            key = f"{slug}/{section_id}"
            fig = re.search(r"<figcaption>([^<]+)", sec.get("content", ""))
            caption = fig.group(1).strip() if fig else ""
            meta = assign.get(slug, {}).get(section_id, {})
            prompt = PROMPTS.get(key, "")
            no = len(rows) + 1
            flat_name = f"{slug}__{section_id}.jpg"
            rows.append([
                no,
                flat_name,
                f"assets/images/photos/{slug}/{section_id}.jpg",
                titles.get(slug, ""),
                caption,
                prompt,
                f"{prompt}. {SUFFIX}" if prompt else "",
                "생성완료" if no <= DONE_UNTIL else "",
                "",
            ])

    flat_names = [r[1] for r in rows]
    if len(flat_names) != len(set(flat_names)):
        dupes = [n for n in flat_names if flat_names.count(n) > 1]
        raise SystemExit(f"파일명 중복: {sorted(set(dupes))}")

    wb = Workbook()
    ws = wb.active
    ws.title = "이미지생성목록"
    headers = ["No", "파일명", "전체 프롬프트"]
    ws.append(headers)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    done_fill = PatternFill("solid", fgColor="E2EFDA")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for row in rows:
        no, flat_name, _deploy, _title, _caption, _prompt, full_prompt, status, _done = row
        ws.append([no, flat_name, full_prompt])

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        is_done = r_idx - 1 <= DONE_UNTIL
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if is_done:
                cell.fill = done_fill

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 90
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24

    g = wb.create_sheet("안내")
    guide = [
        ["항목", "내용"],
        ["총 이미지", f"{len(rows)}장"],
        ["사용법", "No 순서대로 생성 → 파일명 열 그대로 한 폴더에 JPG 저장"],
        ["완료 후", "폴더 경로만 알려주시면 사이트 교체·배포 처리"],
        ["파일명 규칙", "{글slug}__{섹션ID}.jpg (70장 전부 고유)"],
        ["권장 크기", "1200 x 675 px, 16:9, JPG"],
        ["1~12번", "초록색 = 이미 생성한 항목 (파일명만 맞춰 리네임)"],
    ]
    for line in guide:
        g.append(line)
    g.column_dimensions["A"].width = 14
    g.column_dimensions["B"].width = 72
    for cell in g[1]:
        cell.font = Font(name="Arial", bold=True)

    out = ROOT / "image-generation-list.xlsx"
    wb.save(out)
    print(out)


if __name__ == "__main__":
    main()